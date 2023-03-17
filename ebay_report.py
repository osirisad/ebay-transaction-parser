import csv
import os
import shutil
import openpyxl
import itertools
from openpyxl.styles import Alignment
from openpyxl import Workbook
from datetime import datetime
import sys

# Loop through all the CSV files in the import directory
import_folder = './import/'
for filename in os.listdir(import_folder):
    if filename.endswith('.csv'):
        src_file = import_folder + filename
        
        account_type = ''
        with open(src_file, encoding="utf-8-sig") as f:
            
            #get account type
            acc_reader = csv.reader(f)            
            for i in range(8):
                next(acc_reader)
            row = next(acc_reader)
            account_type = row[1]          
        
        start_line = 12
        # Open the CSV file and extract the buyer information               
        with open(src_file, encoding="utf-8-sig") as f:        
            reader = csv.DictReader(itertools.islice(f, start_line-1, None))
            transactions = [row for row in reader]

        # Extract the shipping label costs
        shipping_labels = {}
        for t in transactions:
            if t['Type'] == 'Shipping label':
                shipping_labels[t['Order number']] = t['Gross transaction amount']
                
        other_fees = {}
        for t in transactions:
            if t['Type'] == 'Other fee':
                other_fees[t['Order number']] = t['Gross transaction amount']                                

        # Extract the buyer information and match it with the shipping label cost                       
        buyer_info = [('eBay', account_type, t['Order number'], t['Item title'], 0.0, t['Item subtotal'], t['Quantity'], t['Shipping and handling'], shipping_labels.get(t['Order number'], 0.0), 0.0, 
                       float(t['Final Value Fee - fixed']) + float(t['Final Value Fee - variable']), t['International fee'], other_fees.get(t['Order number'], 0.0), 0.0, t['Buyer name'], t['Buyer username'], 
                       t['Transaction creation date']) for t in transactions if t['Type'] == 'Order']
        
        # Open the Excel file or create it if it doesn't exist
        xlsx_file = 'master_orders.xlsx'

        # Check if the order ID already exists in the file
        order_ids = set()
        try:
            if os.path.isfile(xlsx_file):
                workbook = openpyxl.load_workbook(xlsx_file)
                sheet = workbook["orders"]
                for row in sheet.iter_rows(values_only=True):
                    order_ids.add(row[2])
        except:
            pass

        # Write the new transactions to the file
        for (platform, acc_type, order_id, item_title, my_cost, purchase_price, qty, shipping_charged, actual_shipping_cost, insertion_fees, final_value_fee, 
             international_fee, other_fee, net_profit, buyer_name, buyer_username, date_sold) in buyer_info:
            if order_id not in order_ids:
            
                row_index = sheet.max_row                
                #row_index = max((b.row for b in sheet['B'] if b.value is not None)) 
                
                #convert numbers                                
                purchase_price = float(purchase_price)
                qty = float(qty)
                shipping_charged = float(shipping_charged)
                
                if actual_shipping_cost != '':
                    actual_shipping_cost = float(actual_shipping_cost)
                else:
                    actual_shipping_cost = 0
                
                final_value_fee = float(final_value_fee)
                
                date_sold = datetime.strptime(date_sold, "%b %d, %Y")
                
                if international_fee != '--':
                    international_fee = float(international_fee)
                else:
                    international_fee = 0.0
                    
                other_fee = float(other_fee)                    
        
                #add row
                sheet.append([platform, acc_type, order_id, item_title, my_cost, purchase_price, qty, shipping_charged, actual_shipping_cost, insertion_fees, final_value_fee, 
                              international_fee, other_fee, net_profit, buyer_name, buyer_username, date_sold])
                
                #formatting
                acc_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                
                sheet.cell(row=row_index+1, column=5).number_format = acc_format #my_cost
                sheet.cell(row=row_index+1, column=6).number_format = acc_format #purchase_price
                sheet.cell(row=row_index+1, column=8).number_format = acc_format #shipping_charged
                sheet.cell(row=row_index+1, column=9).number_format = acc_format #actual_shipping_cost
                sheet.cell(row=row_index+1, column=10).number_format = acc_format #insertion_fees
                sheet.cell(row=row_index+1, column=11).number_format = acc_format #final_value_fee
                sheet.cell(row=row_index+1, column=12).number_format = acc_format #international_fee
                sheet.cell(row=row_index+1, column=13).number_format = acc_format #other_fee
                sheet.cell(row=row_index+1, column=14).number_format = acc_format #net_profit
                
                sheet.cell(row=row_index+1, column=17).number_format = 'm/d/yyyy'
                
                #formulas               
                sheet.cell(row=row_index+1, column=14, value='=((F{0}+H{0})+(I{0}+J{0}+K{0}+L{0}+M{0}))+E{0}'.format(row_index+1)).number_format = acc_format  #Net Profit
                
                order_ids.add(order_id)                             

        # Save the changes to the Excel file
        workbook.save(xlsx_file)

        # Move the processed CSV file to the archive subfolder
        dst_folder = './archive/'
        shutil.move(src_file, dst_folder + filename)
